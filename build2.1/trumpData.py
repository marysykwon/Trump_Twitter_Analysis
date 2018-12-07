from openpyxl import *
from trumpLogic import *

#read excel file
book=load_workbook('Trump.xlsx')
sheet=book.active

class readExcel():

    def readTweet(self, tweetRow):
        sentList=[]
        tweet=sheet.cell(column=2,row=tweetRow).value
        cleanedTweet=cleanT(tweet)
        return cleanedTweet

    def readTweetOnly(self,tweetRow):
        tweet=tweet=sheet.cell(column=2,row=tweetRow).value
        tweet=str(tweet)
        return tweet
    
    def readRetweetAmt(self, retweetrow):
        retweetAmt=sheet.cell(column=4,row=retweetrow).value
        return retweetAmt
    
    def readFavAmt(self, favRow):
        favAmt=sheet.cell(column=5,row=favRow).value
        return favAmt

    def readDate(self, daterow):
        tweetDate=str(sheet.cell(column=3, row=daterow).value)
        rawDate=tweetDate.split(" ")[0]
        return rawDate

    def readTime(self, timerow):
        tweetTime=str(sheet.cell(column=3, row=timerow).value)
        rawTime=re.sub("(:)", "", tweetTime[-8:])
        return int(rawTime)

#sentCount[0] = Positive Count
#ratioCount[0]= Positive Ratio
#sentcount[1] = Negative Count
#ratioCount[1] = Negative Ratio
#sentCount[2] = Neutral Count
#ratioCount[2] = Neutral Ratio
#sentiment
def wordSent(sentCount,ratioCount):
    keywordSent=[]
    keywordSent.extend([sentCount[0],ratioCount[0],sentCount[1],ratioCount[1],sentCount[2],ratioCount[2]])
    return keywordSent
